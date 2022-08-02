from bs4 import BeautifulSoup
import datetime
import dateparser
import requests
import pandas as pd
import simplekml
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formula.translate import Translator
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import time
import os
from arcgis.gis import *
from arcgis.gis import GIS
import shutil
from pathlib import Path
import tkinter
from tkinter import *
from Dashapp import *
from threading import Thread
from rain_visual_funcs import *
import random
import warnings



def render_page(url, mode):
    """
    how to webscrape the data
    :param url:
    :param mode:if choose selinium, activate this function
    :return: the soup object for final data acquisiton
    """
    if mode == 'selenium':
        try:
            driver = webdriver.Chrome('C:/Program Files (x86)/chromedriver')
            driver.get(url)
            time.sleep(5)
            r = driver.page_source
            driver.quit()
            soup = BeautifulSoup(r, 'html.parser')
            return soup
        except KeyError:
            print('-----------------------Warning--------------------------')
            print('Check if the Chrome driver version is matching with your browser version')
            print('Resume to the backend mode')
            page = requests.get(url)
            soup = BeautifulSoup(page.content, 'html.parser')
            return soup
    else:
        page = requests.get(url)
        soup = BeautifulSoup(page.content, 'html.parser')
        return soup

# Function to convert string to datetime format
def convert(date_time_in):
    """
    convert str to datatime object
    :param date_time_in: 
    :return: 
    """""
    warnings.filterwarnings(
        "ignore",
        message="The localize method is no longer necessary, as this time zone supports the fold attribute",
    )
    datetime_str = dateparser.parse(date_time_in)
    return datetime_str

# Function to calculate date range
def daterange(start_date, end_date):
    """
    generate a list of all dates within the range
    :param start_date:
    :param end_date:
    :return: generator for dates
    """
    for n in range(int((end_date - start_date).days) + 1):
        yield start_date + datetime.timedelta(n)

# Function to collect data for one day
def fetch_one_day(station, date, paccumchoice, savefolder, mode):
    """
    obtaining one day's rain data
    :param station:
    :param date:
    :param paccumchoice:
    :param savefolder:
    :param mode:
    :return:
    """
    # # debug line
    # station = 'Kcaburli4'
    # date = endDate.strftime("%Y-%m-%d")
    paccumchoice = 'yes'

    url = 'https://www.wunderground.com/dashboard/pws/' + station.upper() + '/table/' + date + '/' + date + '/daily'
    print('fetching page', url)
    soup = render_page(url, mode)

    table_heads = soup.select('table.desktop-table.history-table thead tr th')
    table_rows = soup.select('table.desktop-table.history-table tbody tr')

    data_all_rows = []
    for row in table_rows:
        data_row = []
        all_cells = row.select("td")
        for cell in all_cells:
            try:
                data_row.append(cell.text.replace(u'\xa0°', u' ').strip())
            except:
                data_row.append(np.nan)
        data_all_rows.append(data_row)

    head_names = []
    error_message = ''

    for head in table_heads:
        head_names.append(head.text)

    df_data = pd.DataFrame(data_all_rows)
    if len(df_data) != 0:
        df_data.columns = head_names
        df_data_export = df_data.copy()
        df_data_export['Precip. Accum.'] = df_data_export['Precip. Accum.'].str.extract('(\d*\.\d+|\d+)',expand=False).astype(float)
        df_data_export['Precip. Rate.'] = df_data_export['Precip. Rate.'].str.extract('(\d*\.\d+|\d+)',expand=False).astype(float)
        df_data_export['Time'] = date + ' '+ df_data_export['Time']
        df_data['Time'] = date + ' '+ df_data['Time']
        df_data_export = df_data_export[['Time', 'Precip. Rate.', 'Precip. Accum.']].replace('--', np.NAN)
        df_data_export = df_data_export.rename(columns = {'Time':"datetime", "Precip. Rate.":"prate", 'Precip. Accum.':"paccum"})
        df_data_export.reset_index(drop=True, inplace=True)
        # df_data_export.to_csv(savefolder + '/'+station + '.csv', index=False)
    else:
        df_data_export = pd.DataFrame(columns=["datetime", "prate", "paccum"])
        error_message = f'There is no data available on {date} for site {station}'

    return df_data_export,error_message,df_data

def make_directory(savefolder, subfolder):
    """
    creating directory
    :param savefolder:
    :param subfolder:
    :return:
    """
    d = pathlib.PurePath(savefolder, subfolder)
    p = pathlib.Path(d).mkdir(parents=True, exist_ok=True)
    return d

# Function to Collect Rain Data for one station, every day from start date to end date

def collect_all_days(station, start_date, end_date, paccumchoice, savefolder,download_or_append,oldfilefolder=None,mode=None):
    '''
    This function obtains all RG data from the website and return a list of rain gauge names that are not available
    on the webiste

    :param station: station name from the list
    :param start_date: start date from the list
    :param end_date: end date from the list
    :param paccumchoice: may need to used in the future
    :param savefolder: export csv files to
    :return: a list of rain gauge names that are not available
    on the webiste
    '''
    dfs = [] #data table only contains the rain data
    NA_station = [] #stations that the script failed to collect any data
    dfs_all = [] #full data table from the website

    print(f'-------------Aquatrack will Use Mode {mode} ------------\n')
    error_log =[]
    eval_results = pd.DataFrame(columns=['sitename','Num of Gaps over 15 mins', 'Gaps Location'] )


    for single_date in daterange(start_date, end_date):
        time.sleep(0.5)
        try:
            df_single,daily_error,df_single_all = fetch_one_day(station, single_date.strftime("%Y-%m-%d"), paccumchoice,savefolder, mode)
        # df_single.to_csv(savefolder + '/' + station +'.csv', index=False)
            dfs.append(df_single)
            error_log.append(daily_error)
            dfs_all.append(df_single_all)

        except KeyError:
            site_error = f'Oops, The data for Station {station} is not available at {single_date}, please check the website and consider changing the date range or just skipping this station.'
            error_log.append(site_error)
        try:
            df_alldays = pd.concat(dfs)
            dfs_alldays_full = pd.concat(dfs_all)
            # dfs_alldays_full.to_csv(str(full_table_dir) + '/' + station + '_full_table.csv', index=False)
            # df_old = read_exisitng_rg_data(oldfilefolder, station, file_type='rain_csv')
            # df_alldays = pd.concat([df_alldays, df_old]).drop_duplicates()
            df_alldays['datetime'] = pd.to_datetime(df_alldays['datetime'])
            df_alldays = df_alldays.sort_values(by='datetime')
            df_alldays['datetime'] = df_alldays['datetime'].dt.strftime('%Y-%m-%d %I:%M %p')

            # df_fulltale_old = read_exisitng_rg_data(oldfilefolder, station, file_type='full_table')
            # dfs_alldays_full = pd.concat([dfs_alldays_full, df_fulltale_old]).drop_duplicates()
            dfs_alldays_full['Time'] = pd.to_datetime(dfs_alldays_full['Time'])
            dfs_alldays_full = dfs_alldays_full.sort_values(by='Time')
        except ValueError:
            print(f'No data has been collected for {station}')
            NA_station.append(station)
            continue

    if download_or_append==2:
        print('------------------Appending data------------')

        df_old = read_exisitng_rg_data(oldfilefolder, station, file_type='rain_csv')
        df_alldays = pd.concat([df_alldays, df_old]).drop_duplicates()
        df_alldays['datetime']= pd.to_datetime(df_alldays['datetime'])
        df_alldays = df_alldays.sort_values(by='datetime')
        df_alldays['datetime'] = df_alldays['datetime'].dt.strftime('%Y-%m-%d %I:%M %p')

        df_fulltale_old = read_exisitng_rg_data(oldfilefolder,station,file_type='full_table')
        dfs_alldays_full = pd.concat([dfs_alldays_full, df_fulltale_old]).drop_duplicates()
        dfs_alldays_full['Time'] = pd.to_datetime(dfs_alldays_full['Time'])
        dfs_alldays_full = dfs_alldays_full.sort_values(by='Time')

        ## Make directories for each type of download
    rain_csv_dir = make_directory(savefolder, 'rain_csv')
    full_table_dir = make_directory(savefolder, 'full_table')

    df_alldays.to_csv(str(rain_csv_dir) + '/' + station + '.csv', index=False)
    dfs_alldays_full.to_csv(str(full_table_dir) + '/' + station + '_full_table.csv', index=False)

    if len(rain_eval(df_alldays, station)):
        eval_results = rain_eval(df_alldays, station)
    else:
        eval_results = 'NO DATA'
    #     eval_results.to_csv(savefolder + '/' + station + 'rg_evaluation.csv', index=False)

    error_log = [i for i in error_log if i!='']
    return NA_station, error_log,eval_results

def rain_eval(df_alldays, station):
    '''
    To use this function, you need to put all your downloaded rain data to seperae folder
    :param folder:
    :return:
    '''

    data = df_alldays
    if 'datetime' in data.columns:
        data_no_nan = data.dropna().copy()
        data_no_nan['datetime'] = pd.to_datetime(data_no_nan['datetime'])
        data_no_nan['gap'] = data_no_nan['datetime'].diff().dt.seconds > 15*60
        gap_sum = data_no_nan['gap'].sum()
        sitename = station
        gap_loc = data_no_nan.loc[data_no_nan['gap']==True, 'datetime']
        eval_results = [sitename, gap_sum, gap_loc]
        return eval_results
    else:
        print(f'the datetime column is missing for {station}.')

def fill_excel(stationName,exceltemp, savefolder):
    wb = load_workbook(exceltemp)
    ws = wb.active
    excel_dir = make_directory(savefolder,'processd_excel')
    csvfolder = pathlib.PurePath(savefolder, 'rain_csv')

    #  import station data
    if os.path.isfile(str(csvfolder)+'/'+ stationName + '.csv'):
        a = pd.read_csv(str(csvfolder)+'/'+ stationName + '.csv', header=0)
        # Remove extra headers
        a = a.drop(a[a['datetime'] == 'datetime'].index)
        a['datetime'] = pd.to_datetime(a['datetime'], format='%Y-%m-%d %I:%M %p')
        a.reset_index(inplace=True)
        a.paccum = a.paccum.astype(float)
        a.drop(['index'], axis=1, inplace=True) # this is probably no longer needed as we have updated the webscraping code. Now we won't see bunch of 'datatime' rows
        #  Convert excel rows
        rows = dataframe_to_rows(a, index=False)
        # fill the excel table

        for r_idx, row in enumerate(rows, 6):  # starts at 6 as you want to skip the first 5 rows
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx>10:
                ws[f'D{r_idx}'] = Translator(ws['D10'].value, origin='D10').translate_formula(f'D{r_idx}')
                ws[f'E{r_idx}'] = Translator(ws['E10'].value, origin='E10').translate_formula(f'E{r_idx}')
                ws[f'F{r_idx}'] = Translator(ws['F10'].value, origin='F10').translate_formula(f'F{r_idx}')
                ws[f'H{r_idx}'] = Translator(ws['H10'].value, origin='H10').translate_formula(f'H{r_idx}')
                ws[f'I{r_idx}'] = Translator(ws['I10'].value, origin='I10').translate_formula(f'I{r_idx}')
                ws[f'J{r_idx}'] = Translator(ws['J10'].value, origin='J10').translate_formula(f'J{r_idx}')
                ws[f'K{r_idx}'] = Translator(ws['K10'].value, origin='K10').translate_formula(f'K{r_idx}')
                ws[f'L{r_idx}'] = Translator(ws['L10'].value, origin='L10').translate_formula(f'L{r_idx}')

        ws.auto_filter.ref = f'H7:L{ws.max_row}'
        ws.auto_filter.add_filter_column(0, ['Keep'], blank = False)
        ws.auto_filter.add_sort_condition(f'H7:H{ws.max_row}')
        wb.save(str(excel_dir)+'/'+ stationName+"_processed.xlsx")




# Making kml file
def kml_making(df_coordinate_all, savefolder, stationlist):
    kml = simplekml.Kml()
    for n in range(len(df_coordinate_all)):
        Name = df_coordinate_all.iloc[n,0]
        Coords = [(float(df_coordinate_all.loc[n,'Longitude (Degree)']), float(df_coordinate_all.loc[n,'Latitude (Degree)']))]
        kml.newpoint(name = Name, coords=Coords) # lon, lat, optional height
        print(Coords)
    kml.save(savefolder+'/'+os.path.split(stationlist)[1].split('.')[0]+'.kml')


def download_img(project_id, savepath):
    user = 'nkwan_v_and_a'
    password = '1000sewerRAT'
    gis = GIS('https://v-and-a.maps.arcgis.com', user, password)
    un = gis.properties.user.username
    print(f' Logged in as {un}')

    id = 'a73f76ec01cd40268b704a423e5d8fc5'  # using FMgbd_download as a test
    # project_id = '20-0340' #project photos to download
    gdb = gis.content.get(id)
    assert gdb is not None, "No matching project for this ID. Please Check the GIS layer ID."

    attac_tbl = gdb.tables[2].query(out_sr=4326, as_df=True)  # access the table information
    attac_tbl_select = attac_tbl.loc[attac_tbl['site_id_3'].str.contains(project_id, na=False),
                       :]  # only contains the selcted project
    attac_tbl_select.reset_index(inplace=True)
    ids = attac_tbl_select['OBJECTID'].tolist()
    # savefolder = 'C:\\Users\\PXie\\Documents\\Python Projects\\wunderground\\attachement download\\photosbysites'
    # dest_dir = 'C:\\Users\\PXie\\Documents\\Python Projects\\wunderground\\attachement download\\all_photos'

    savefolder = os.path.join(savepath, 'by_sites')
    dest_dir = os.path.join(savepath, 'all_photos')

    if not os.path.exists(savefolder):
        os.makedirs(savefolder)

    if not os.path.exists(dest_dir):
        os.makedirs(dest_dir)

    imgs = []

    for i in range(len(attac_tbl_select)):
        site_id = attac_tbl_select.loc[i, 'site_id_3']
        oid = attac_tbl_select.loc[i, 'OBJECTID'].tolist()
        print(f'downloading image for {site_id}')
        try:
            img = gdb.tables[2].attachments.download(oid=oid, save_path=savefolder + f'\\{site_id}')
            for root, dirs, files in os.walk(savefolder + f'\\{site_id}'):
                file_dir = glob.glob(root + '\\*.jpg')
                imgs.append(file_dir)
                for file_name in file_dir:
                    if file_name:
                        print(root)
                        print(len(files))
                        print(f'Currently processing {file_name}')
                        print(f'-------------------------------------------------------')
                        print('\n')
                        seperator = '\\'
                        if not site_id in file_name.split('\\')[-1]:
                            newname = seperator.join(file_name.split('\\')[:-1]) + '\\' + site_id + '_' + \
                                      file_name.split('\\')[-1]
                            os.rename(file_name, newname)
                            shutil.copy(newname, dest_dir)
        except:
            print('a photo was failed to upload')


def get_install_kmz_func(savefolder, project_id):
    user = 'nkwan_v_and_a'
    password = '1000sewerRAT'
    gis = GIS('https://v-and-a.maps.arcgis.com', user, password)
    un = gis.properties.user.username
    print(f' Logged in as {un}')

    id = 'a73f76ec01cd40268b704a423e5d8fc5'  # using FMgbd_download as a test
    project_id = project_id  # project photos to download
    gdb = gis.content.get(id)
    assert gdb is not None, "No matching project for this ID. Please Check the GIS layer ID."

    attac_tbl = gdb.layers[0].query(out_sr=4326, as_df=True)  # access the table information
    attac_tbl_select = attac_tbl.loc[attac_tbl['project_no'].str.contains(project_id, na=False),
                       :].reset_index(drop=True)  # only contains the selcted project
    attac_tbl_select = attac_tbl_select.dropna(subset=['install_date']).reset_index(drop=True)
    kml = simplekml.Kml()
    for n in range(len(attac_tbl_select)):
        Name = attac_tbl_select.loc[n,'site_name']
        print(Name)
        Coords = [(float(attac_tbl_select.loc[n,'SHAPE'].x), float(attac_tbl_select.loc[n,'SHAPE'].y))]
        Dia_final = None
        if  attac_tbl_select.loc[n,'removal_diameter'] >1:
            Dia_final = attac_tbl_select.loc[n,'removal_diameter']
        elif attac_tbl_select.loc[n,'install_diameter']>1:
            Dia_final = attac_tbl_select.loc[n, 'install_diameter']
        else:
            Dia_final = attac_tbl_select.loc[n, 'expected_diam']


        if '.5' in str(Dia_final):
            Dia_final = (Dia_final)
        else:
            Dia_final = int(Dia_final)

        if attac_tbl_select.loc[n, 'direction']:
            direction = attac_tbl_select.loc[n, 'direction'].upper()
        else:
            direction =''
        in_out = attac_tbl_select.loc[n, 'in_out']
        Name_final = Name + f'_({Dia_final})_{direction} {in_out}'
        kml.newpoint(name = Name_final, coords=Coords) # lon, lat, optional height

    kml.newpoint(name=Name, coords=Coords)  # lon, lat, optional height

    kml.save(savefolder + '/' + project_id + '_fmgdb_installed.kml')
    print('the KMZ file is created')

def get_tidal_func(tidal_list):
    tidal_pd = pd.read_csv(tidal_list, header=None)
    tidal_pd = tidal_pd.dropna()

    saveto = os.path.dirname(tidal_list)
    product = 'water_level'

    for i in range(len(tidal_pd)):
        station_name = tidal_pd.iloc[i, 0]
        start_date = convert(tidal_pd.iloc[i, 1])
        end_date = convert(tidal_pd.iloc[i, 2])
        url_csv = f'https://api.tidesandcurrents.noaa.gov/api/prod/datagetter?product={product}' \
                  f'&application=NOS.COOPS.TAC.WL&begin_date={start_date}&end_date={end_date}&datum=MLLW&station={station_name}&time_zone=LST&units=english&format=csv'
        req_csv = requests.get(url_csv)
        with open(f"{saveto}/{station_name}_tidal.csv", 'w', encoding=req_csv.encoding) as csvFile:
            csvFile.write(req_csv.text, )

        print(f'The tidal for {station_name} is downloaded successfully.')

def get_idf_func(coordinate_list):
    coordinate_pd = pd.read_csv(coordinate_list)
    saveto = os.path.dirname(coordinate_list)
    idf_dir= make_directory(saveto,'IDFs')

    for i in range(len(coordinate_pd)):
        station_name = coordinate_pd.iloc[i,0]
        lon = str(coordinate_pd.iloc[i,2])
        lat = str(coordinate_pd.iloc[i,1])
        url_csv = f"https://hdsc.nws.noaa.gov/cgi-bin/hdsc/new/fe_text_mean.csv?lat={lat}6&lon={lon}&data=depth&units=english&series=pds&selAddr=Burlingame, California, USA&selElevNum=511.04&selElevSym=ft&selStaName=-"
        req_csv = requests.get(url_csv)

        with open(f"{str(idf_dir)}/{station_name}_idf.csv", 'w', encoding=req_csv.encoding) as csvFile:
            csvFile.write(req_csv.text, )

        print(f'The IDF data for {station_name} is downloaded successfully.')

def run_func(stationlist,choice,download_or_append,savefolder, exceltemp):
    paccumchoice = "Yes"
    print(f'-----downlaod or append = {download_or_append}------ ')
    file = open(stationlist, "r")
    all_coordinate_dict = {}
    # Repeat for each station in the .csv file
    error_log = []  # collect all error message from the run
    eval_results = pd.DataFrame(columns=['sitename', 'Num of Gaps over 15 mins', 'Gaps Location'])

    if download_or_append == 2:
        oldfilefolder = tkinter.filedialog.askdirectory(
            title='Choose the folder where the original downloaded files are saved or select the save to folder for downloading rain data with multiple periods')
    else:
        oldfilefolder = savefolder

    driver_path = Path('C:/Program Files (x86)/chromedriver.exe')
    if not driver_path.is_file():
        print('-----------------------Warning--------------------------')
        print('Can not find Chrome driver at: C:/Program Files (x86)')
        print('Change to the backend mode')
        mode = 'backend'
    elif choice == 2 and driver_path.is_file():
        mode = 'selenium'
    else:
        mode = 'backend'

    for line in file:
        time.sleep(0.3)
        # Let's split the line into an array called "fields" using the "," as a separator:
        fields = line.split(",")
        # and let's extract the data:
        stationName = fields[0]
        startDate_str = fields[1]
        endDate_str = fields[2]
        startDate = convert(startDate_str)
        endDate = convert(endDate_str)
        print("Get " + stationName + " from: " + startDate_str + " to: " + endDate_str)

        # collect rain data and the rain gauge coordination


        NA_station, site_error_log, rgdata_eval = collect_all_days(stationName, startDate, endDate, paccumchoice,
                                                                   savefolder, download_or_append,oldfilefolder,mode=mode)

        eval_results.loc[eval_results.shape[0]] = rgdata_eval

        error_log.append(site_error_log)
        if stationName not in NA_station:
            # coordinates = coordinate(stationName, startDate.strftime("%Y-%m-%d"))
            # all_coordinate_dict[stationName] = coordinates
            # df_coordinate_all = pd.DataFrame(all_coordinate_dict).T.rename(
            #     columns={1: 'Latitude (Degree)', 0: 'Longitude (Degree)'})
            fill_excel(stationName, exceltemp, savefolder)
        else:
            print(f'The Station {stationName} is not not available on the website')
            print('\n\n')

    # It is good practice to close the file at the end to free up resources
    file.close()

    # save the coordination file

    # save the error log
    df_error_log = pd.DataFrame(error_log)
    df_error_log.dropna(axis='columns', inplace=True)
    df_error_log = df_error_log.T
    df_error_log.to_csv(
        savefolder + '/' + os.path.split(stationlist)[1].split('.')[0] + '_error_log.csv', index=False, header=False)

    eval_results.to_csv(savefolder + '/' + 'rg_evaluation.csv', index=False)

def coordination_func(rg_list):
    coordinate_pd = pd.read_csv(rg_list, header=None, index_col=False)
    saveto = os.path.dirname(rg_list)
    coordinate_pd.iloc[:, 1] = pd.to_datetime(coordinate_pd.iloc[:, 1])
    df_coordinate_all = pd.DataFrame(columns=['StationName', 'Latitude (Degree)', 'Longitude (Degree)','Elevation (ft)'])
    coordinate_list = []
    for i in range(len(coordinate_pd)):
        station = coordinate_pd.iloc[i, 0]
        date = coordinate_pd.iloc[i, 1].strftime("%Y-%m-%d")
        url = 'https://www.wunderground.com/dashboard/pws/' + station.upper() + '/table/' + date + '/' + date + '/daily'
        print('fetching coordinate')
        page = requests.get(url)
        soup = BeautifulSoup(page.content, 'html.parser')
        ## getting the elevation info
        ele_tag = soup.body.find("div", attrs={'class': 'sub-heading'})
        span_tags = ele_tag.find_all('span', recursive=False)
        str_to_remove = re.findall(r"(?<=<)[^><]*(?=>)", str(span_tags[0]))
        sub_heading_list = str(span_tags[0]).replace('Elev', '').replace('ft', '').replace('<', '').replace('>', '')
        for i in str_to_remove:
            sub_heading_list = sub_heading_list.replace(i, '')
        ele = int(sub_heading_list.split(',')[0])
        print(f'The Elevation value for Station {station} is: {ele}')

        ## finding hidden longitude and latitude
        try:
            test = soup.find_all("script", attrs={'id': 'app-root-state'})
            test_content = test[0].contents[0]
            pattern_lon = re.compile(r"lon&q;:(.*?),&q;")
            pattern_lat = re.compile(r"lat&q;:(.*?),&q;")
            lon = (pattern_lon.findall(test_content)[0])
            lat = (pattern_lat.findall(test_content)[0])
            print(f'The longitude value for Station {station} is: {lon}')
            print(f'The latitude value for Station {station} is: {lat}')
        except:  # if the backend returns nothing, then mannually open the web page and click the detail button
            try:
                path = "C:\Program Files (x86)\chromedriver.exe"
                driver = webdriver.Chrome(path)
                driver.get('url')

                time.sleep(2)
                button = driver.find_element_by_xpath(
                    '//*[@id="inner-content"]/div[1]/app-dashboard-header/div[2]/div/div[2]/div/lib-pws-info-icon/mat-icon')
                button.click()
                divs = driver.find_element_by_class_name('cdk-overlay-container')
                a = divs.text
                b = list(filter(lambda x: 'Latitude / Longitude' in x, a.split('\n')))[0]

                Num = re.findall(r'[0-9]+', b)
                lat = int(Num[0]) + int(Num[1]) * 0.001
                lon = -int(Num[2]) - int(Num[3]) * 0.001

                print(f'The longitude value for Station {station} is: {lon}')
                print(f'The latitude value for Station {station} is: {lat}')
                time.sleep(1)
                driver.quit()
            except:
                print('-----------------------Warning-------------------------------------')
                print(f'site {station} is not available. Please, double check the website.')
                print('-------------------------------------------------------------------')
                continue
        coord_persite = [station, lat, lon, ele]
        df_coordinate_all.loc[df_coordinate_all.shape[0]] = coord_persite

    df_coordinate_all.to_csv(saveto + '/' + os.path.split(rg_list)[1].split('.')[0] + '_coordinates.csv', index=False)
    return df_coordinate_all, saveto, rg_list


def open_dashapp(port):
    url = f'http://127.0.0.1:{port}/'
    try:
        chrome_options = Options()
        chrome_options.add_experimental_option("detach", True)
        driver = webdriver.Chrome('C:/Program Files (x86)/chromedriver',options=chrome_options)
        driver.get(url)
    except KeyError:
        print('-----------------------Warning--------------------------')
        print('Check if the Chrome driver version is matching with your browser version')
        print('You can still access the app by typing the address into your webbrowser')

def create_rain_qaqc():
    template_path = tkinter.filedialog.askopenfilename(title='Select the Rain QAQC WU&VA Template excel file ',
                                                       filetypes = (('Excel files','*.xlsx'),))
    rg_location_path = tkinter.filedialog.askopenfilename(title='Select the Rain Site Coordination ',filetypes = (('CSV files','*.csv'),))
    processed_file_path = tkinter.filedialog.askdirectory(title = 'Choose the folder where the processes excel files are saved')
    saveto_folder = tkinter.filedialog.askdirectory(title = 'Choose the folder where the QAQC file will be saved into')
    processed_xlsx_paths = glob.glob(processed_file_path + "/*.xlsx")
    rg_location = pd.read_csv(rg_location_path)

    excelApp = client.gencache.EnsureDispatch("Excel.Application")
    wb = excelApp.Workbooks.Open(template_path)
    excelApp.Visible = False
    ws = wb.Worksheets('Para')
    project_num = input('Input the Project Number:  ')
    owner = input('Input the Owner:  ')
    location = input('Input the Location:  ')
    start_date = input('Input the Start Date, i.e. 2020-01-30:  ')
    num_sites = len(processed_xlsx_paths)

    ws.Range('B3').Value = project_num
    ws.Range('B4').Value = owner
    ws.Range('B5').Value = location
    ws.Range('B6').Value = start_date

    for idx, val in enumerate(processed_xlsx_paths):
        ws = wb.Worksheets('Para')
        site_name = Path(val).stem.split('_')[0]
        if site_name in list(rg_location['StationName']):
            lat = rg_location.loc[rg_location['StationName'] == site_name, 'Latitude (Degree)'].item()
            lon = rg_location.loc[rg_location['StationName'] == site_name, 'Longitude (Degree)'].item()
            ele = rg_location.loc[rg_location['StationName'] == site_name, 'Elevation (ft)'].item()
        else:
            lat = 'NA'
            lon = 'NA'
            ele = 'NA'
        ws.Range(f'C{12 + idx}').Value = site_name
        ws.Range(f'D{12 + idx}').Value = lat
        ws.Range(f'E{12 + idx}').Value = lon
        ws.Range(f'F{12 + idx}').Value = ele
        ws.Range(f'G{12 + idx}').Formula = f"='{idx+1}'!F3"

        ## working on the datasheet
        ws = wb.Worksheets(f'{idx + 1}')
        print(f'------------Working on {site_name}--------------')
        process_wb = excelApp.Workbooks.Open(val)
        ws_xlsx = process_wb.Worksheets('ather17')  # current sheet
        df = pd.DataFrame(ws_xlsx.UsedRange())
        df_filtered = df.iloc[6:, 7:].copy()
        df_filtered = df_filtered.dropna()
        df_filtered.columns = ['1', '2', '3', '4', '5']
        df_filtered = df_filtered[df_filtered['4'] == 'Keep']
        process_wb.Close()

        ws.Range(ws.Cells(7, 7),  # Cell to start the "paste"
                 ws.Cells(7 + df_filtered.shape[0] - 1,
                          7 + df_filtered.shape[1] - 1)
                 ).Value = df_filtered.values

        ws.Range(f"A{7}:F{7 + df_filtered.shape[0] - 1}").FillDown()
        ws.Range('F3').Formula = f'=MAX(F6:F{7 + df_filtered.shape[0] - 1})'

    savefile = os.path.join(saveto_folder, f'{project_num} {location} QAQC Rain VA&WU ({num_sites} sites).xlsx')
    saveFileNoSlash = os.path.normpath(savefile)
    print(f'the spread sheet has been saved into {saveFileNoSlash}')
    wb.SaveAs(saveFileNoSlash)
    wb.Close()

def coordinate():
    """
    This function generates a kml file and a csv file for the locations of the rgs
    :return:
    """
    rg_list = tkinter.filedialog.askopenfilename(title='Select the rain gauge list file (the same you used for downloading rain data',
                                                 filetypes = (('CSV files','*.csv'),))
    df_coordinate_all, saveto, rg_list = coordination_func(rg_list)
    kml_making(df_coordinate_all, saveto, rg_list)


def get_idf():
    """
    webscrapting the idfs
    :return:
    """
    coordinate_list = tkinter.filedialog.askopenfilename(
        title='Select the coordinate list file', filetypes= (('CSV files','*.csv'),))
    get_idf_func(coordinate_list)

def get_tidal():
    """
    webscraping the tidal data
    :return:
    """
    tidal_list = tkinter.filedialog.askopenfilename(
        title='Select the tidal list file', filetypes= (('CSV files','*.csv'),))
    get_tidal_func(tidal_list)


def get_timelist(flow_df):
    Time_list = list(set(flow_df['Datetime'].dt.date.tolist()))
    Time_list.sort()
    return Time_list


def initial_dash():
    excel_path = tkinter.filedialog.askopenfilename(
                    title='Open a flow qaqc excel xlsx file',
                    initialdir='/',
                    filetypes=(('Excel files', '*.xlsx'),))
    flow_df,rain_df = read_flow_rain_from_excel(excel_path)
    port = random.randint(1000,9999)
    flow_df = flow_df.dropna(subset=['Flow'])
    try:
        # FFT_App(flow_df, rain_df,port)
        # open_dashapp(port)

        t1 = Thread(target=FFT_App, args=(excel_path,flow_df, rain_df, port))
        # t2 = Thread(target=open_dashapp, args=(port, ))
        #
        t1.start()
        # time.sleep(3)
        # t2.start()
    except:
        raise Exception("something is wrong")


def rain_visual():
    excel_path = tkinter.filedialog.askopenfilename(title='Select the Rain QAQC WU&VA excel file ',filetypes = (('Excel files','*.xlsx'),))
    # boundary_path = tkinter.filedialog.askopenfilename(title='Select the boundary shape file ',filetypes = (('Shape files','*.shp'),))
    outputfolder = tkinter.filedialog.askdirectory(title = 'Choose the folder to save the exported figs and gif')

    para_df, num_site, site_list = get_para_df(excel_path)
    Data_df, interval = get_rain_df(excel_path, site_list)

    start_date = input('Please choose the start time of study period, i.e 2022-03-26 or 2022-03-26 06:30:  ')
    end_date = input('Please choose the end time of study period, i.e 2022-03-26 or 2022-03-26 06:30:  ')
    Data_df = Data_df.loc[start_date:end_date]

    start_date = change_time_format(start_date)
    end_date = change_time_format(end_date)
    sub_folder = interval + '_' + start_date + '_to_' + end_date
    outputfolder_new = make_directory(outputfolder, subfolder=sub_folder)

    # import boundary and flow basin polygon
    print('Input the longitude of the western point of the bounding box:  ')
    xmin = degree_to_decimal('W')
    print('Input the longitude of eastern point of the bounding box:  ')
    xmax = degree_to_decimal('W')
    print('Input the latitude of the southern point of the bounding box:  ')
    ymin = degree_to_decimal('N')
    print('Input the latitude northern point of the bounding box:  ')
    ymax = degree_to_decimal('N')

    nx = int(input('Number of cells in longitudinal direction:  '))
    ny = int(input('Number of cells in latitudinal direction:  '))

    export_fig_animation(nx, ny, xmin, xmax,ymin,ymax, Data_df, para_df, outputfolder_new)
    print('The figures and animation are being produced')
    print('--------------------------------------------')


def read_exisitng_rg_data(savefolder,NA_Station,file_type):
    '''
    this function read the pre-downloadded data
    :param savefolder:
    :param NA_Station:
    :param file_type:
    :return:
    '''
    NA_Station = NA_Station
    savefolder = savefolder
    try:
        if file_type == 'rain_csv':
            csvfile = savefolder+'\\rain_csv\\'+NA_Station+'.csv'
            df = pd.read_csv(csvfile)
        if file_type =='full_table':
            csvfile = savefolder + '\\full_table\\' + NA_Station + '_full_table.csv'
            df = pd.read_csv(csvfile)
    except FileNotFoundError:
        # print("No Existing Data Found. Will Create a Brand New Rain File")
        return None
    return df
